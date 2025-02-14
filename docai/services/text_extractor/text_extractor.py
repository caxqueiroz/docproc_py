from typing import Optional
import PyPDF2
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image
import pytesseract
from pathlib import Path
import io

class TextExtractor:
    @staticmethod
    def _convert_pdf_page_to_image(page) -> Image.Image:
        """Convert a PDF page to a PIL Image"""
        # Convert PDF page to image
        try:
            import fitz  # PyMuPDF
            pix = page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))  # 300 DPI
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            return img
        except ImportError:
            # Fallback to pdf2image if PyMuPDF is not available
            from pdf2image import convert_from_bytes
            import io
            
            # Get page as bytes
            writer = PyPDF2.PdfWriter()
            writer.add_page(page)
            pdf_bytes = io.BytesIO()
            writer.write(pdf_bytes)
            pdf_bytes.seek(0)
            
            # Convert to image
            images = convert_from_bytes(pdf_bytes.getvalue(), dpi=300)
            return images[0] if images else None

    @staticmethod
    def extract_from_pdf(file_path: Path) -> str:
        # First try normal text extraction
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        
        # If no text was extracted, try OCR
        if not text.strip():
            try:
                import fitz  # PyMuPDF
                pdf_document = fitz.open(file_path)
                for page_num in range(len(pdf_document)):
                    page = pdf_document[page_num]
                    img = TextExtractor._convert_pdf_page_to_image(page)
                    if img:
                        page_text = pytesseract.image_to_string(img)
                        if page_text:
                            text += page_text + "\n"
                pdf_document.close()
            except ImportError:
                # Fallback to pdf2image if PyMuPDF is not available
                from pdf2image import convert_from_path
                images = convert_from_path(file_path, dpi=300)
                for img in images:
                    page_text = pytesseract.image_to_string(img)
                    if page_text:
                        text += page_text + "\n"
        
        return text.strip()

    @staticmethod
    def extract_from_docx(file_path: Path) -> str:
        doc = DocxDocument(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])

    @staticmethod
    def extract_from_xlsx(file_path: Path) -> str:
        wb = load_workbook(filename=file_path, read_only=True)
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                row_text = " ".join(str(cell.value) for cell in row if cell.value is not None)
                if row_text:
                    text.append(row_text)
        return "\n".join(text)

    @staticmethod
    def extract_from_image(file_path: Path) -> str:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image)

    def extract_text(self, file_path: Path) -> Optional[str]:
        file_extension = file_path.suffix.lower()
        
        extractors = {
            '.pdf': self.extract_from_pdf,
            '.docx': self.extract_from_docx,
            '.xlsx': self.extract_from_xlsx,
            '.png': self.extract_from_image,
            '.jpg': self.extract_from_image,
            '.jpeg': self.extract_from_image,
        }
        
        extractor = extractors.get(file_extension)
        if not extractor:
            raise ValueError(f"Unsupported file type: {file_extension}")
            
        return extractor(file_path)
